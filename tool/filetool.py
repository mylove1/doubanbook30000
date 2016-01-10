__author__ = 'hunterhug'
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
import re
# 找出文件夹下所有html后缀的文件
def listfiles(rootdir, prefix='.html'):
	file = []
	for parent,dirnames,filenames in os.walk(rootdir):
		if parent == rootdir:
			for filename in filenames:
				if filename.endswith(prefix):
					file.append(filename)
			return file
		else:
			pass

def rlistfiles(rootdir, prefix='.html'):
	file = []
	for parent,dirnames,filenames in os.walk(rootdir):
		for filename in filenames:
			if filename.endswith(prefix):
				#file.append(filename)
				print(str(parent)+'/'+filename)
			else:
				pass
	return file

def writeexcel(path, content, sheetname='标签抓取表'):
	wb=Workbook()
	sheet=wb.create_sheet(0,sheetname)
	row = 1
	col = 1
	for i in content:
		for j in i:
			sheet.cell(row=row,column=col).value = j
			col = col + 1
		col = 1
		row =row + 1
	wb.save(path)
	print("保存数据成功！")

def readexcel(path):
	excelcontent = []
	wb2=load_workbook(path)
	sheetnames = wb2.get_sheet_names()
	ws=wb2.get_sheet_by_name(sheetnames[0])
	row=ws.get_highest_row()
	col=ws.get_highest_column()
	# print("列数: ",ws.get_highest_column())
	# print("行数: ",ws.get_highest_row())

	for i in range(0,row):
		rowcontent = []
		for j in range(0,col):
			if ws.rows[i][j].value:
				rowcontent.append(ws.rows[i][j].value)
		excelcontent.append(rowcontent)
	print("读取数据成功！")
	return excelcontent

# 去除标题中的非法字符 (Windows)
def validateTitle(title):
	rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/\:*?"<>|'
	new_title = re.sub(rstr, "", title)
	return new_title


if __name__=='__main__':
	rlistfiles('../book')